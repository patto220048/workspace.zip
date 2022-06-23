import openpyxl
import os 
from PIL import Image
import numpy as np


wb= openpyxl.load_workbook('aloha.xlsx')

ws= wb['Sheet']

path = './mnist_mini_1000/'
files = os.listdir(path)

Num4Files = 0



kptae=''
for file in files:
    if 'num4' in file or 'num9' in file:
        if 'num4' in file:
            kotae=4
        else:
            kotae=9
        
        img=Image.open(path+file)
        imgArry = np.array(img)
        for i in range (9):
            for j in range(9):
                ws.cell(row=1+i, column=1+j+9*Num4Files).value = imgArry[i][j]
        ws.cell(row=10,column=1+9*Num4Files).value=kotae
        Num4Files+=1




wb.save('aloha.xlsx')